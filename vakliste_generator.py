"""
 Auto-create a roster spreadsheet from a list of experiments and on-duty observers.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Color, PatternFill, Font, Alignment
import itertools
from session import Session
from observer import Observer


def generate_workbook():
    """Main function to generate the Vakliste workbook/spreadsheet for presentation."""

    debug = True

    # initialise: #
    wb = Workbook()
    ws = wb.active

    #########################################################################################
    # define colour and colour list:
    # name = color # compliment
    red = "00FF0000"  # cyan
    purple = "00800080"  # 008000
    blue = "003366FF"  # FFCC33
    teal = "00008080"  # 800000
    cyan = "0000FFFF"  # red
    lavender = "009999FF"  # FFFF99
    # need to think about what colours would actually look nice, these above are just place-holders

    # list of all colours we would like in the spreadsheet:
    colour_list = [red, purple, blue, teal, cyan, lavender]
    #########################################################################################

    # create list of observer objects:
    obs_list = get_observer_list_from_file("observers_onduty.txt", colour_list)

    # create list of experiment objects:
    exp_list = get_exp_list_from_file("experiments_nn_ns.txt")

    # sort the list into ascending d.o.y. (not really necessary?)
    exp_list = sorted(exp_list, key=lambda x: (x.doy, x.ut))

    if debug:
        print("------------------------------------------------------")
        print(f"Number of Experiments this month: {len(exp_list)}")

    # distribute the experiments equally between the observers:
    schedule = distribute_shifts(obs_list, exp_list)

    # flip the dictionary so that given an experiment, can get an observer
    reverse_lookup = create_reverse_lookup(schedule)

    if debug:
        print("Number of shifts per person on-duty:")
        for staff, shifts in schedule.items():
            print(staff.name, ": ", len(shifts))
        print("------------------------------------------------------")

    if debug:
        print("EXP, TELE, D.O.Y., DATE, NAME, START (UT), SHIFT (LT), DUR, WEEK #, ON-DUTY")
        for exp in exp_list:
            on_duty = reverse_lookup.get(exp.name)
            print(exp.name, exp.tele, exp.doy, exp.get_start_date(), exp.get_name_of_start_day(), exp.ut,
                  exp.get_lt_shift_start(), exp.duration,
                  exp.get_week_num(), on_duty.name)
        print("------------------------------------------------------")

    #
    setup_worksheet(ws, exp_list, reverse_lookup, colour_list)

    # save the file
    wb.save(filename="test.xlsx")
    #########################################################################################


def setup_worksheet(ws, exp_list, reverse_lookup, colour_list):

    # fill in the info
    populate_worksheet(ws, exp_list, reverse_lookup, colour_list)

    # add some colours and borders
    style_worksheet(ws, colour_list)
    #########################################################################################


def populate_worksheet(ws, exp_list, reverse_lookup, colour_list):
    """Fill in the values of the cells."""

    [red, purple, blue, teal, cyan, lavender] = colour_list

    # fill in the first column
    set_cell(ws, 1, 1, "Week", "")
    set_cell(ws, 2, 1, "SESSION", red)
    set_cell(ws, 3, 1, "TELESCOPE", blue)
    set_cell(ws, 4, 1, "DATE", teal)
    set_cell(ws, 5, 1, "DAY", teal)
    set_cell(ws, 6, 1, "d.o.y.", "")
    set_cell(ws, 7, 1, "START (LT)", blue)

    ###############################################
    # configure start points
    i = 2
    j = 8

    # start filling out the heard of the spreadsheet
    for exp in exp_list:

        set_cell(ws, 1, i, f"{exp.get_week_num()}")
        set_cell(ws, 2, i, f"{exp.name}", red)
        set_cell(ws, 3, i, f"{exp.tele}", blue)
        set_cell(ws, 4, i, f"{exp.get_start_date()}", teal)
        set_cell(ws, 5, i, f"{exp.get_name_of_start_day()}", teal)
        set_cell(ws, 6, i, f"{exp.doy}")
        set_cell(ws, 7, i, f"{exp.get_lt_start()}", blue)

        # check if session overflows to following day
        dur_hr, dur_min = exp.get_duration()
        k = 1 + (dur_hr // 24)
        for l in range(k):

            j = j + l
            set_cell(ws, j, 1, (reverse_lookup.get(exp.name)).name, reverse_lookup.get(exp.name).colour)

            # if no overflow
            if k == 1:
                set_cell(ws, j, i, f"{exp.get_lt_shift_start()}-{exp.get_lt_shift_end()}", color=reverse_lookup.get(exp.name).colour)

            # if overflows
            elif k == 2:
                if l == 0:
                    set_cell(ws, j, i, f"{exp.get_lt_shift_start()}-08:00", color=reverse_lookup.get(exp.name).colour)
                elif l == 1:
                    set_cell(ws, j, i, f"08:00-{exp.get_lt_shift_end()}", color=reverse_lookup.get(exp.name).colour)

        i += 1
        j += 1
        #########################################################################################


def style_worksheet(ws, colour_list):
    """Fill in the border & colours."""

    [red, purple, blue, teal, cyan, lavender] = colour_list

    #
    alternating_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for cell in row:
            if row[0].row % 2 == 0:  # Apply fill to even rows
                cell.fill = alternating_fill

    #
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(bottom=Side(style='thick'))

    #
    for col in ws["A"]:
        col.border = Border(left=Side(style='thick'), right=Side(style='thick'))

    #
    ws[f"A{ws.max_row}"].border = Border(left=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'))

    #
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.border = Border(bottom=Side(style='thick'), top=Side(style='thick'))

    #
    ws["A1"].border = Border(left=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'))

    #
    for row in ws.iter_rows(min_row=7, max_row=7):
        for cell in row:
            cell.border = Border(bottom=Side(style='thick'))

    #
    for row in ws.iter_cols(ws.max_column):
        for cell in row:
            cell.border = Border(right=Side(style='thick'))

    #
    ws["A7"].border = Border(left=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'))

    #
    for cc in ws.columns:
        ws.column_dimensions[get_column_letter(cc[0].column)].width = 15

    #
    for row in ws.iter_rows(min_row=1, max_row=7):
        for cell in row:
            if cell.value == "Nn":
                for r in range(2, 8):
                    ws.cell(row=r, column=cell.column).fill = PatternFill(patternType='lightUp', fgColor=Color(lavender))
            if cell.value == "Ns":
                for r in range(2, 8):
                    ws.cell(row=r, column=cell.column).fill  = PatternFill(patternType='lightUp', fgColor=Color(teal))

    #
    ws.cell(row=ws.max_row, column=ws.max_column).border = Border(bottom=Side(style='thick'), right=Side(style='thick'))
    ws.cell(row=1, column=ws.max_column).border = Border(bottom=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'))
    ws.cell(row=7, column=ws.max_column).border = Border(bottom=Side(style='thick'), right=Side(style='thick'))
    #########################################################################################


def set_cell(ws, i, j, value, color=""):
    """Set the cell value & style."""

    ws.cell(row=i, column=j).value = value
    ws.cell(row=i, column=j).alignment = Alignment(shrinkToFit=False, horizontal='center', vertical='center')

    if not color:
         ws.cell(row=i, column=j).font = Font(bold=True)
    else:
        ws.cell(row=i, column=j).font = Font(color=color, bold=True)
    #########################################################################################


def create_reverse_lookup(schedule):
    """Get experiment given who is on duty."""

    reverse_lookup = {}
    for staff, shifts in schedule.items():
        for shift in shifts:
            reverse_lookup[shift] = staff
    return reverse_lookup
    #########################################################################################


def distribute_shifts(staff_list, exp_list):
    """Associate an experiment with an on-duty observer."""

    schedule = {staff: [] for staff in staff_list}
    staff_cycle = itertools.cycle(staff_list)

    st_tmp = None
    doy_tmp = None
    current_staff = next(staff_cycle)

    for exp in exp_list:

        if exp.ut != st_tmp or exp.doy != doy_tmp:
            current_staff = next(staff_cycle)

        st_tmp = exp.ut
        doy_tmp = exp.doy

        schedule[current_staff].append(exp.name)
    return schedule
    #########################################################################################


def get_observer_list_from_file(filename, colour_list):
    """Create list of observer objects."""

    observers = []
    i = 0
    file = open(filename, 'r')
    lines = file.readlines()
    for line in lines:
        color = colour_list[i]
        who = line.strip()
        i += 1
        obs = Observer(who, color)
        observers.append(obs)
    return observers
    #########################################################################################


def get_exp_list_from_file(filename):
    """Create list of experiment objects."""

    exp_list = []
    file = open(filename, 'r')
    lines = file.readlines()
    for line in lines:
        lc = line.split()
        exp = Session(lc[0], lc[1], lc[2], lc[3], lc[4])
        # name, telescopes, doy_start, ut_start, duration
        exp_list.append(exp)
    file.close()
    return exp_list
    #########################################################################################


if __name__ == '__main__':

    generate_workbook()
    #########################################################################################
